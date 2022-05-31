# ************************************************************************
# ****  TXpira v1.0 / 2019 - Maximiliano Urtubey - urtubey@gmail.com  ****
# ************************************************************************

import sys

import openpyxl
libro = openpyxl.load_workbook('expire.xlsx')
hoja = libro.active

tot_fil = hoja.max_row
tot_col = hoja.max_column

tabla = list()
renglon = list()
for x in range(1, tot_fil+1):
    for y in range(1, tot_col): ### < no se le agrega el 1 para no importar los días vigentes de la lista de Excel
        renglon.append(hoja.cell(row=x, column=y).value)
    tabla.append(renglon)
    renglon = []
    x += 1

col_cbar = [] # crea lista de la columna de código de barras completos
for rt in tabla:
    #int_f = int(rt[0])
    #col_cbar.append(int_f)
    col_cbar.append(rt[0])
    
col_cbar4 = [] # crea lista de los últimos cuatro dígitos con la columna de código de barras
for f in tabla:
    str_f = str(f[0])
    str_4 = str_f[-4:]
    col_cbar4.append(str_4)
    
col_fv = [] # crea lista con la columna de fecha de vencimiento
for f in tabla:
    int_f1 = int(f[-2])
    str_f1 = str(int_f1)
    col_fv.append(str_f1)

m = 0 # agrega un '0' a los ítems de la tabla de 5 dígitos
for it in col_fv:
    sit = str(it)
    if len(sit) < 6:
        rem = '0' + sit
        col_fv.pop(m)
        col_fv.insert(m, rem)
    m += 1

u = 0 # convierte a string los números de fecha de vencimiento restantes
for it in col_fv:
    if type(it) != 'str':
        xit = str(it)
        col_fv.pop(u)
        col_fv.insert(u, xit)
    u += 1

i = 0 # reemplaza la primera columna de la tabla por la lista col_cbar
for fila in tabla:
    fila.pop(0)
    fila.insert(0, col_cbar[i])
    i += 1

i = 0 # reemplaza los valores de fecha de vencimiento por valores corregidos lista col_fv
for fila in tabla:
    fila.pop(-2)
    fila.insert(-1, col_fv[i])
    i += 1
    
for fila in tabla: # convierte los las unidades de flotantes a enteros
    un = int(fila[-1])
    fila[-1] = un

import smtplib
def env_mail(listta):
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.ehlo()
    server.starttls()
    server.ehlo()
    
    server.login('direccionEnvio@gmail.com', 'codigode16digitos')
    subject = 'Ver próximos vencimientos'
    body = ''
    for line in (listta):
        body += "%s %s %s %s %s %d %s %d\n" %(line[1], line[2], '| venc:', line[3], '| cant.', line[4], '| días rest:', line[5])
    msg = f'Subject: {subject}\n\n{body}'.encode('utf-8') #sin el ".encode('utf-8')" tira error
    server.sendmail(
        'direccionEnvio@gmail.com', # from
        'direccionDestino@gmail.com', # to
        msg)
    print('Se envió un recordatorio por e-mail')
    server.quit()

# <<<<<<<<<<<<<<<<<<<<<<  imprime la tabla por primera vez >>>>>>>>>>>>>>>>>>>>>>>>>>>

import copy
from tabulate import tabulate 
import fechas
new_tabla = copy.deepcopy(tabla)
fechas.fecha_form(new_tabla)# agrega una columna con los días vigentes a 'new_tabla' /// 'tabla' no tiene la columna con los días vigentes
import orden
tablord = orden.sort(new_tabla)# ordena 'new_tabla' 
print(tabulate(tablord, headers=['Cod barra', 'nombre', 'marca', 'venc.', 'cant', 'validz']))

nota = ''
pven = []
for vig in tablord:
    if vig[-1] <= 30:
        nota = 'verdd'
        pven.append(vig)

if nota == 'verdd':
    print()
    print('Hay items que están prontos a vencer:')
    for gg in pven:
        print(f'{gg[1]} {gg[2]} venc.: {gg[3]} ,{gg[5]} días rest, cant.: {gg[4]}')
    print()
    print('¿Desea enviar un recordatorio por mail? (s/n)')
    while True:
        envm = input('>>> ')
        if envm == 's':
            env_mail(pven)
            print(f'Recordatorio enviado')
            break
        elif envm == 'n':
            print('no se envió recordatorio alguno')
            break
        else:
            print()
            print("'s' ó 'n' tecleá de nuevo")


def f_frlp(cb, lstolst):
    for row in lstolst:
        srow = str(row[0])
        srow_4 = srow[-4:]
        if cb == srow_4:
            fv = row[-2]
            ct = row[-1]
            return srow_4, fv, ct, row


def find_index(cb, fven, lstolst):
    ind = 0
    conj = []
    posc = []
    for row in lstolst:
        srow = str(row[0])
        srow_4 = srow[-4:]
        if cb == srow_4:
            conj.append(row[-2])# agrega a 'conj' la fecha de vencimiento
            posc.append(ind)# agrega a 'posc' el nro ind
        ind += 1
    if fven in conj:
        var = 'existente'
        pla = conj.index(fven)
        indix = posc[pla]
        return var, conj, posc, indix
    else:
        var = 'inexistente'
        return var
    

#<<<------------------------------------->>>#
while True:

    repet = set([x for x in col_cbar4 if col_cbar4.count(x) > 1])# lista con códigos repetidos en base a 4 dígitos
    #print(repet)
    
    cb = input(f'''
        ingrese 4 últim. Cod Barras:
        salir (x) ó guardar (g)
        >>> ''').lower()
    if cb == 'x':
        sys.exit()
            
    if cb == 'g':
        save = input(f'''
        desea guardar (s)/(n):
        >> ''').lower()
        if save == 's':
            import up_2_date_xlsx
            up_2_date_xlsx.actualizar('expire.xlsx', tablord)# guarda el archivo ya ordenado
            print('''
        *** GUARDADO ***''')
            continue
        elif save == 'n':
            print('----')
            continue
        else:
            print()
            print(" 's' o 'n' perejil !!!")
            continue
    
    if cb.isalpha() or len(cb) != 4:# verifica que sean sólo cuatro dígitos
        print()
        print('///sólo los cuatro últimos dígitos///')
        print('///      intente nuevamente       ///')
        continue
    
    if cb not in col_cbar4:#(A) no existente en la lista.
        print()
        print('nuevo producto...')
        cbar = int(input('ingrese el código de barras completo: > '))
        stricbar = str(cbar)
        if len(stricbar) != 13:
            print()
            print('revise el código -vuelva a intentar-')
            continue
        
        if cb != stricbar[-4:]:
            print()
            print('No coincide con el nro. ingresados anteriormente:')
            print(f'cuatro últimos dígitos: .....{cb}')
            print('                             ↓↓↓')
            print(f'código completo: {cbar}')
            pat = input(f'''
        ¿el código completo es el correcto?
        (s)/(n) >>>''')
            if pat == 's':
                pass
            else:
                continue
                
        desc = input('producto? > ')
        brnd = input('marca? > ')
        fven = input('ingrese fecha de vencimiento > ')
        cant = int(input('cantidad? > '))
        nfil = [cbar, desc, brnd, fven, cant]
        tabla.append(nfil)
        scbar = str(cbar)
        scbar4 = scbar[-4:]
        col_cbar4.append(scbar4)
        
        new_tabla = copy.deepcopy(tabla)                                                                     ######
        fechas.fecha_form(new_tabla)# agrega columna con días de vigencia                                    ######
        print()                                                                                              ######
        print(f'Lista actualizada (nuevo): {desc} {brnd}')
        tablord = orden.sort(new_tabla)# ordena 'new_tabla' en 'tablord' imprime fecha actual                ######
        print(tabulate(tablord, headers=['Cod barra', 'nombre', 'marca', 'venc.', 'cant', 'validz']))        ######
            
    else:
        qua = 0 # suma las cantidades si hay elementos repetidos
        for x in tabla:
            sx = str(x[0])
            sx_4 = sx[-4:]
            if cb == sx_4 and len(col_cbar4) != len(set(col_cbar4)):
                qua = qua + x[-1]

        g = f_frlp(cb, tabla)# verifica si ya existe el ítem en la tabla.
        if cb == g[0]:
            print(f"producto existente: {g[3][1]} {g[3][2]} cant: {qua} ")
            elec = input(f'''
            Almacenar --> A
            Consumir  --> C
            > ''').lower()
            if elec == 'a':
                fve = input(f'ingrese fecha vencimiento -ddmmaa- > ' )
                dd = int(fve[:2])
                mm = int(fve[2:4])
                aa = int(fve[-2:])
                if dd > 31 or mm > 12:
                    print('fecha/formato incorrecto')
                    print(fve)
                    continue
                
                cuant = int(input(f'cantidad? >'))
                
                if cb not in repet and cb == g[0] and fve != g[1]:#(B) almacenar existente UNICO con nueva fecha vencimiento. ( <?> WORKING)
                    print(f'fv del existente: {g[1]}')
                    nfil = [g[3][0], g[3][1], g[3][2], fve, cuant]
                    tabla.append(nfil)
                    col_cbar4.append(cb)
                    
                    new_tabla = copy.deepcopy(tabla)                                                                     ######
                    fechas.fecha_form(new_tabla)# agrega columna con días de vigencia                                    ######
                    print()                                                                                              ######
                    print('Lista actualizada (almacenar nuevo vencimiento): ')
                    tablord = orden.sort(new_tabla)# ordena 'new_tabla' en 'tablord' imprime fecha actual                ######
                    print(tabulate(tablord, headers=['Cod barra', 'nombre', 'marca', 'venc.', 'cant', 'validz']))        ######
                                                    
                elif cb not in repet and cb == g[0] and fve == g[1]:#(C) almacenar existente UNICO con misma fecha vencimiento. (WORKING)
                    print(g[3][-1])
                    nnc = g[3][-1] + cuant
                    g[3].pop()
                    g[3].append(nnc)
                    
                    new_tabla = copy.deepcopy(tabla)                                                                     ######
                    fechas.fecha_form(new_tabla)# agrega columna con días de vigencia                                    ######
                    print()                                                                                              ######
                    print('Lista actualizada (almacenar mismo vencimiento):')
                    tablord = orden.sort(new_tabla)# ordena 'new_tabla' en 'tablord' imprime fecha actual                ######
                    print(tabulate(tablord, headers=['Cod barra', 'nombre', 'marca', 'venc.', 'cant', 'validz']))        ######
                                                        
                elif cb in repet:# almacenar repetido con misma o distinta fecha de vencimiento.                                (WORKING)
                    indx = find_index(cb, fve, tabla)# < funciona cuando existe o no la fecha de vencimiento
                    if indx[0] == 'existente':
                        nnc = tabla[indx[3]][-1] + cuant
                        tabla[indx[3]].pop()
                        tabla[indx[3]].append(nnc)
                        
                        new_tabla = copy.deepcopy(tabla)                                                                     ######
                        fechas.fecha_form(new_tabla)# agrega columna con días de vigencia                                    ######
                        print()                    
                        print('Lista actualizada (REPETIDO mismo vencimiento):')
                        tablord = orden.sort(new_tabla)# ordena 'new_tabla' en 'tablord' imprime fecha actual                ######
                        print(tabulate(tablord, headers=['Cod barra', 'nombre', 'marca', 'venc.', 'cant', 'validz']))        ######
                                            
                    elif indx == 'inexistente':# <<< no es indx[0] porque es el único que devuelve (return var)
                        nrow = [g[3][0], g[3][1], g[3][2], fve, cuant]
                        tabla.append(nrow)
                        
                        new_tabla = copy.deepcopy(tabla)                                                                     ######
                        fechas.fecha_form(new_tabla)# agrega columna con días de vigencia                                    ######
                        print()                    
                        print('Lista actualizada (REPETIDO distinto vencimiento):')
                        tablord = orden.sort(new_tabla)# ordena 'new_tabla' en 'tablord' imprime fecha actual                ######
                        print(tabulate(tablord, headers=['Cod barra', 'nombre', 'marca', 'venc.', 'cant', 'validz']))        ######
                                                                                            
            elif elec == 'c':
                g = f_frlp(cb, tabla)
                if cb not in repet and cb == g[0] and g[2] > 1:# no existen items iguales con != fecha vencim y hay + de 1.    (WORKING)
                    #print('único mayor a 1')
                    nnc = g[2] - 1
                    g[3].pop()
                    g[3].append(nnc)
                    
                    new_tabla = copy.deepcopy(tabla)                                                                     ######
                    fechas.fecha_form(new_tabla)# agrega columna con días de vigencia                                    ######
                    print()                    
                    print(f'Lista actualizada (se consumió: {g[3][1]} {g[3][2]})')
                    tablord = orden.sort(new_tabla)# ordena 'new_tabla' en 'tablord' imprime fecha actual                ######
                    print(tabulate(tablord, headers=['Cod barra', 'nombre', 'marca', 'venc.', 'cant', 'validz']))        ######
                                                        
                elif cb not in repet and cb == g[0] and g[2] == 1:# existen un ítem únicamente.                                 (WORKING)
                    print()
                    print(f'NO QUEDA: {g[3][1]} {g[3][2]}')
                    nnc = g[2] - 1
                    g[3].pop()
                    g[3].append(nnc)
                    lista_n = [n for n in tabla if n[-1] != 0]
                    tabla = lista_n[:]
                    col_cbar4.remove(cb)
                    
                    new_tabla = copy.deepcopy(tabla)                                                                     ######
                    fechas.fecha_form(new_tabla)# agrega columna con días de vigencia                                    ######
                    print()                    
                    print(f'Lista actualizada (se consumió: {g[3][1]} {g[3][2]})')
                    tablord = orden.sort(new_tabla)# ordena 'new_tabla' en 'tablord' imprime fecha actual                ######
                    print(tabulate(tablord, headers=['Cod barra', 'nombre', 'marca', 'venc.', 'cant', 'validz']))        ######
                                
                elif cb in repet:# está incluido en repet (hay + de 1 con distintas fechas de venc)                             (WORKING)
                    ven = []
                    for row in tabla:
                        srow = str(row[0])
                        srow_4 = srow[-4:]
                        if cb == srow_4:
                            ven.append(row[-2])# agrega a 'ven' la fecha de vencimiento
                    print(ven)
                    print()
                    sfven = input(f'ingrese vencimiento -ddmmaa- (existen + de 1) > ')
                    if sfven not in ven:
                        print()
                        print('no existe esa fecha de vencimiento')
                        continue

                    indx = find_index(cb, sfven, tabla)
                    #print(f'vencimiento: {tabla[indx[3]][-2]}')
                    #print(f'cantidad: {tabla[indx[3]][-1]}')
                    if sfven == tabla[indx[3]][-2] and tabla[indx[3]][-1] > 1:# existen items iguales con != fecha vencim + de 1. 
                        nnc = tabla[indx[3]][-1] - 1
                        tabla[indx[3]].pop()
                        tabla[indx[3]].append(nnc)
                        
                        new_tabla = copy.deepcopy(tabla)                                                                     ######
                        fechas.fecha_form(new_tabla)# agrega columna con días de vigencia                                    ######
                        print()
                        print('Lista actualizada (mismo ítem varias fechas de vencimiento):')
                        tablord = orden.sort(new_tabla)# ordena 'new_tabla' en 'tablord' imprime fecha actual                ######
                        print(tabulate(tablord, headers=['Cod barra', 'nombre', 'marca', 'venc.', 'cant', 'validz']))        ######
                                  
                    elif sfven== tabla[indx[3]][-2] and tabla[indx[3]][-1] == 1:# existen items iguales con != fecha vencim = a 1.
                        print()
                        print(f'no queda: {g[3][1]} {g[3][2]} con fecha de venc. {tabla[indx[3]][-2]}')
                        tabla.pop(indx[3])# elimina la fila del último elemento
                        #col_cbar4.remove(cb)  ESTE ES EL PUTO ERROR POR EL CUAL NO APARECÍA EN LA LISTA DE REPETIDOS
                        
                        new_tabla = copy.deepcopy(tabla)                                                                     ######
                        fechas.fecha_form(new_tabla)# agrega columna con días de vigencia                                    ######
                        print()
                        print('Lista actualizada (consumir duplicado = 0):')
                        tablord = orden.sort(new_tabla)# ordena 'new_tabla' en 'tablord' imprime fecha actual                ######
                        print(tabulate(tablord, headers=['Cod barra', 'nombre', 'marca', 'venc.', 'cant', 'validz']))        ######
            else:
                print('ingrese A / C únicamente')
                continue
