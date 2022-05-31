import openpyxl

def actualizar(libro, tabla):# (libro) es el archivo de excel, (tabla) nombre de lista 2d
    
    lib_act = openpyxl.load_workbook(libro)
    pag = lib_act.active
    r = str(pag.max_row)
    co = pag.max_column
    rco = co + 64 # caracter del número máximo de columnas (65 es 'A' en ASCII)
    letra = chr(rco)
    n = letra + r
    for row in pag["A1":n]:# rango desde A1 hasta la última columna/fila
        for cell in row:
            cell.value = None
    for xr, row in enumerate(tabla):# loop para guardar los valores en celdas
            for yc, val in enumerate(row):
                pag.cell(xr + 1, yc + 1, val)
    lib_act.save(libro)
    
pass